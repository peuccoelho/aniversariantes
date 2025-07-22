from dotenv import load_dotenv
import os
import pandas as pd
from datetime import datetime
import calendar
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import Font, Alignment

load_dotenv()

arquivo_excel = os.getenv("ARQUIVO_EXCEL")

if not arquivo_excel:
    print("Variável ARQUIVO_EXCEL não encontrada no .env.")
    exit()

# padronizar e converter datas
def converter_data(valor):
    if pd.isna(valor):
        return pd.NaT

    valor_str = str(valor).strip()


    # Se vier só dia e mês (DDMM)
    if valor_str.isdigit() and len(valor_str) == 4:
        dia = valor_str[:2]
        mes = valor_str[2:4]
        ano = "1900"  # Ano padrão
        valor_str = f"{dia}/{mes}/{ano}"

    # Se vier no formato DD/MM (sem ano)
    elif len(valor_str) == 5 and valor_str[2] in ["/", "."]:
        dia = valor_str[:2]
        mes = valor_str[3:5]
        ano = "1900"
        valor_str = f"{dia}/{mes}/{ano}"

    # Se vier dia, mês e ano (DDMMYY ou DDMMYYYY)
    elif valor_str.isdigit() and len(valor_str) == 6:
        dia = valor_str[:2]
        mes = valor_str[2:4]
        ano = valor_str[4:]
        if int(ano) <= 30:
            ano = "20" + ano
        else:
            ano = "19" + ano
        valor_str = f"{dia}/{mes}/{ano}"

    formatos = ["%d/%m/%Y", "%d.%m.%Y", "%d%m%Y", "%d/%m/%y", "%d.%m.%y", "%d%m%y"]
    for f in formatos:
        try:
            return datetime.strptime(valor_str, f)
        except:
            continue

    try:
        return pd.to_datetime(valor_str, dayfirst=True, errors='coerce')
    except:
        return pd.NaT

df = pd.read_excel(arquivo_excel, dtype=str)
df.columns = df.columns.str.strip().str.upper()

# conversão de datas
df["DATA DE NASCIMENTO"] = df["DATA DE NASCIMENTO"].apply(converter_data)
df = df[df["DATA DE NASCIMENTO"].notna()]

mes_escolhido = int(input("Digite o número do mês (1 a 12): "))
meses_pt = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]
nome_mes = meses_pt[mes_escolhido]

import numpy as np
# filtro de aniversariantes (aceita datas sem ano, usando ano padrão 1900)
aniversariantes = df[df['DATA DE NASCIMENTO'].apply(lambda x: isinstance(x, (datetime, np.datetime64)) and x.month == mes_escolhido)].copy()

if aniversariantes.empty:
    print(f"Nenhuma pessoa faz aniversário em {nome_mes}.")
    exit()

print(f"{len(aniversariantes)} aniversariante(s) encontrado(s) para {nome_mes}.")

# ordem por dia
aniversariantes['DIA'] = aniversariantes['DATA DE NASCIMENTO'].dt.day
aniversariantes = aniversariantes.sort_values(by='DIA')
aniversariantes.drop(columns=['DIA'], inplace=True)

# dados para exportação
colunas_saida = ['DATA DE NASCIMENTO', 'GH', 'NOME COMPLETO', 'SETOR']
aniversariantes_saida = aniversariantes[colunas_saida].copy()
aniversariantes_saida['DATA DE NASCIMENTO'] = aniversariantes_saida['DATA DE NASCIMENTO'].dt.strftime('%d/%m/%Y')

nome_arquivo_excel = f'aniversariantes_mes_{mes_escolhido:02}.xlsx'
nome_aba = f"Aniversariantes_{nome_mes}"

# Excel
with pd.ExcelWriter(nome_arquivo_excel, engine='openpyxl') as writer:
    aniversariantes_saida.to_excel(writer, sheet_name=nome_aba, index=False, startrow=2)

wb = load_workbook(nome_arquivo_excel)
ws = wb[nome_aba]
ws.merge_cells('A1:D1')
ws['A1'] = f"Aniversariantes do mês {nome_mes}"
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')
wb.save(nome_arquivo_excel)

# PDF
nome_arquivo_pdf = f'aniversariantes_mes_{mes_escolhido:02}.pdf'
doc = SimpleDocTemplate(nome_arquivo_pdf, pagesize=A4)
styles = getSampleStyleSheet()
titulo_style = ParagraphStyle(
    name='CenterTitle',
    parent=styles['Title'],
    alignment=1,
    fontSize=16,
    spaceAfter=12
)
elementos = []
titulo = f"Aniversariantes do mês {nome_mes}"
elementos.append(Paragraph(titulo, titulo_style))
elementos.append(Spacer(1, 12))

dados_pdf = aniversariantes_saida.copy()
dados_pdf['DATA DE NASCIMENTO'] = pd.to_datetime(
    dados_pdf['DATA DE NASCIMENTO'], format='%d/%m/%Y', errors='coerce'
).dt.strftime('%d/%m')


colunas_pdf = ['DATA DE NASCIMENTO', 'GH', 'NOME COMPLETO', 'SETOR']
dados = [colunas_pdf] + dados_pdf.values.tolist()

Tabela = Table(dados, hAlign='CENTER')
Tabela.setStyle(TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, -1), 10),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
]))
elementos.append(Tabela)
doc.build(elementos)

print(f"Arquivos gerados com sucesso: {nome_arquivo_excel} e {nome_arquivo_pdf}")